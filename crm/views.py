import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import FileResponse, JsonResponse
from django.forms import modelformset_factory
from .models import Order, OrderFile, Client, OrderProduct, Product, Transport
from .forms import BaseOrderProductFormSet, UserRegistrationForm, UserLoginForm, OrderForm, OrderProductForm, OrderFileForm, UserProfileUpdateForm, PasswordChangeForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
import os
from decimal import Decimal
from django.views.decorators.csrf import ensure_csrf_cookie
from django.core.serializers.json import DjangoJSONEncoder
from google.oauth2 import service_account
from googleapiclient.discovery import build

@ensure_csrf_cookie # Этот декоратор гарантирует установку CSRF cookie
def login_page(request):
    """Renders the login page."""
    return render(request, 'login.html')

def login_api(request):
    """Handles the login logic for an API request."""
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return JsonResponse({'success': True, 'redirect_url': '/orders/'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    return JsonResponse({'error': 'Only POST method is allowed'}, status=405)


@ensure_csrf_cookie # Этот декоратор также нужен и для страницы регистрации
def register_page(request):
    """Renders the registration page."""
    return render(request, 'register.html')

def register_api(request):
    """Handles the registration logic for an API request."""
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            login(request, user)
            return JsonResponse({'success': True, 'redirect_url': '/orders/'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    return JsonResponse({'error': 'Only POST method is allowed'}, status=405)


def user_logout(request):
    logout(request)
    return redirect('login_page')


@login_required
def order_list_page(request):
    return render(request, 'orderList.html')

# Функция для выдачи JSON данных
@login_required
@permission_required('crm.can_view_all_orders', raise_exception=True)
def order_list_api(request):
    orders_queryset = Order.objects.select_related('client', 'transport').all().order_by('-created_at')
    orders_list = []
    for order in orders_queryset:
        car_display_number = 'N/A'
        if order.transport:
            # Собираем в список все части, которые не пустые
            parts = [order.transport.car_number, order.transport.trailer_number]
            # Объединяем их через слэш
            car_display_number = "/".join(filter(None, parts)) or 'N/A'

        orders_list.append({
            'id': order.id,
            'date': order.created_at.strftime('%Y-%m-%d'),
            'number': f"ORDER-{order.order_number}",
            'carNumber': car_display_number,
            'grossWeight': f"{order.gross_weight} kg",
            'netWeight': f"{order.netto_weight} kg",
            'price': f"{order.total_amount} {order.currency}",
        })
    return JsonResponse({'orders': orders_list})

@login_required
def order_detail_page(request, order_id):
    """Render the initial HTML page for the order details."""
    # Мы просто передаем ID заказа в шаблон, чтобы JS мог его использовать
    return render(request, 'orderDetails.html', {'order_id': order_id})

# ПРЕДСТАВЛЕНИЕ 2: Отдает JSON-данные и обрабатывает загрузку файлов
@login_required
@permission_required('crm.can_view_all_orders', raise_exception=True)
def order_detail_api(request, order_id):
    """Return details for a single order as JSON and handle file uploads."""
    order = get_object_or_404(Order, id=order_id)

    # Обработка загрузки нового файла
    if request.method == 'POST':
        form = OrderFileForm(request.POST, request.FILES)
        if form.is_valid():
            order_file = form.save(commit=False)
            order_file.order = order
            order_file.save()
            # Возвращаем имя файла в случае успеха
            return JsonResponse({'success': True, 'file_name': order_file.file.name.split('/')[-1]})
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    # Для GET-запроса отдаем детали заказа
    # Собираем информацию о товарах
    products_list = []
    for item in order.order_products.all():
        products_list.append({
            'name': item.product.product_name,
            'quantity': item.quantity,
            'grossWeight': f"{item.gross_weight} kg",
            'amount': f"{item.amount} {order.currency}",
        })

    # Собираем информацию о файлах
    files_list = []
    for item in order.files.all():
        files_list.append({
            'id': item.id,
            'name': item.file.name.split('/')[-1], # Получаем только имя файла
            'url': f'/orders/{order.id}/download/{item.id}/'
        })

    # Собираем все данные в один объект
    data = {
        'client': order.client.name,
        'crossing': order.transition_point,
        'destination': order.destination,
        'sender': order.shipper,
        'receiver': order.consignee,
        'senderAddress': order.shipper_address,
        'receiverAddress': order.consignee_address,
        'carNumber': order.transport.car_number if order.transport else '',
        'carWeight': str(order.transport.car_weight) if order.transport and order.transport.car_weight is not None else '',
        'trailerNumber': order.transport.trailer_number if order.transport else '',
        'trailerWeight': str(order.transport.trailer_weight) if order.transport and order.transport.trailer_weight is not None else '',
        'driver_number': order.transport.driver_number if order.transport else '',
        'loadedWeight': str(order.loaded_weight) if order.loaded_weight is not None else '',
        'currency': order.currency,
        'goods': products_list,
        'files': files_list,
    }

    return JsonResponse(data)

@login_required
@permission_required('crm.can_view_all_orders', raise_exception=True)
def download_file(request, order_id, file_id):
    """Handle file downloads for a specific order."""
    order = get_object_or_404(Order, id=order_id)
    order_file = get_object_or_404(OrderFile, id=file_id, order=order)
    file_path = order_file.file.path
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=os.path.basename(file_path))
    messages.error(request, 'File not found.')
    return redirect('order_detail', order_id=order_id)

@login_required
@permission_required('crm.can_manage_orders', raise_exception=True)
def order_create_page(request):
    """Renders the blank form page for creating a new order."""
    # Передаем в шаблон список всех клиентов и продуктов
    # чтобы JavaScript мог построить из них выпадающие списки
    clients = Client.objects.filter(users__user=request.user)
    products_data = list(Product.objects.values('id', 'product_name', 'price_mult'))

    currency_choices = Order.CURRENCY_CHOICES
    
    return render(request, 'orderCreate.html', {
        'clients': clients,
        'products_json': json.dumps(products_data, cls=DjangoJSONEncoder),
        'currency_choices': currency_choices
    })


def append_to_google_sheet(order_data):
    try:
        # ID вашей Google Таблицы
        SPREADSHEET_ID = '17UoXOEmRJfKXQ8fwaH46jaGK2T4yhXSbQrzm_hK7gUo' # <-- ЗАМЕНИТЕ НА СВОЙ ID

        # Путь к файлу с ключами
        SERVICE_ACCOUNT_FILE = 'credentials.json'
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)

        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        # Формируем строку для добавления
        row_data = [
            order_data['date'],
            order_data['order_number'],
            '',                           
            '',
            order_data['car_info'],
            order_data['product_name'],
            order_data['gross_weight'],
            order_data['net_weight'],
            order_data['total_amount']
        ]

        # Добавляем строку в таблицу на лист с названием 'Sheet1'
        request = sheet.values().append(spreadsheetId=SPREADSHEET_ID,
                                        range="Sheet1!A1",
                                        valueInputOption="USER_ENTERED",
                                        insertDataOption="INSERT_ROWS",
                                        body={"values": [row_data]})
        response = request.execute()
        print(f"--- Google Sheets: Данные добавлены: {response}")

    except Exception as e:
        # В случае ошибки просто выводим ее в консоль, чтобы не ломать основной процесс
        print(f"--- ОШИБКА Google Sheets: {e}")

# ПРЕДСТАВЛЕНИЕ 2: API для создания заказа из JSON-данных
@login_required
@permission_required('crm.can_manage_orders', raise_exception=True)
def order_create_api(request):
    """
    Создает новый заказ, перераспределяет вес и генерирует инвойс.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)

    try:
        data = json.loads(request.body)
        
        if not data.get('client_id') or not data.get('products'):
            return JsonResponse({'error': 'Client and products are required.'}, status=400)
        
        products_input_data = data.get('products', [])

        # --- НАЧАЛО ОБНОВЛЕННОЙ ЛОГИКИ ПЕРЕРАСЧЕТА ВЕСА ---
        if len(products_input_data) >= 2:
            
            for p_data in products_input_data:
                product_model = get_object_or_404(Product, id=p_data['product_id'])
                p_data['price_mult'] = product_model.price_mult
                p_data['gross_weight'] = Decimal(p_data['gross_weight'])
            
            products_input_data.sort(key=lambda p: p['price_mult'])
            
            cheapest_product = products_input_data[0]
            most_expensive_product = products_input_data[-1]
            
            MIN_WEIGHT_THRESHOLD = Decimal('2000.00')
            if most_expensive_product['gross_weight'] >= MIN_WEIGHT_THRESHOLD:
                
                # Расчет базового веса для переброса (как и раньше)
                total_weight_pair = most_expensive_product['gross_weight'] + cheapest_product['gross_weight']
                expensive_ratio = most_expensive_product['gross_weight'] / total_weight_pair if total_weight_pair > 0 else 0
                transfer_factor = Decimal(1.0) - abs(Decimal(0.5) - expensive_ratio)
                base_weight_to_transfer = (Decimal(2) * transfer_factor * Decimal('3500.00')).min(Decimal('3500.00'))

                # --- НОВОЕ ПРАВИЛО: Ограничение в 41% ---
                # Рассчитываем максимальный вес для переброса по новому правилу
                max_transfer_by_percentage = most_expensive_product['gross_weight'] * Decimal('0.41')

                # Выбираем НАИМЕНЬШЕЕ из двух значений: базового расчета и лимита в 41%
                weight_to_transfer = min(base_weight_to_transfer, max_transfer_by_percentage)
                
                # Округляем и финализируем вес для переброса
                weight_to_transfer = round(weight_to_transfer, 2)
                weight_to_transfer = min(weight_to_transfer, most_expensive_product['gross_weight'] - MIN_WEIGHT_THRESHOLD)

                if weight_to_transfer > 0:
                    print(f"--- Перерасчет: Перебрасываем {weight_to_transfer} кг с дорогого на дешевый ---")
                    most_expensive_product['gross_weight'] -= weight_to_transfer
                    cheapest_product['gross_weight'] += weight_to_transfer

        # --- Логика создания/поиска транспорта ---
        transport = None
        car_number = data.get('car_number')
        if car_number:
            transport, created = Transport.objects.get_or_create(
                car_number=car_number,
                defaults={
                    'trailer_number': data.get('trailer_number'),
                    'driver_number': data.get('driver_number'),
                    'car_weight': data.get('car_weight') or None,
                    'trailer_weight': data.get('trailer_weight') or None
                }
            )
            if not created:
                # Обновляем данные, если транспорт уже существовал
                transport.trailer_number = data.get('trailer_number', transport.trailer_number)
                transport.driver_number = data.get('driver_number', transport.driver_number)
                transport.car_weight = data.get('car_weight', transport.car_weight)
                transport.trailer_weight = data.get('trailer_weight', transport.trailer_weight)
                transport.save()
        
        # --- ИСПРАВЛЕНИЕ ЗДЕСЬ: Логика расчета номера заказа ---
        last_order = Order.objects.order_by('-order_number').first()
        new_order_number = (last_order.order_number + 1) if last_order else 1

        # --- Создание объекта заказа с указанием номера ---
        client = get_object_or_404(Client, id=data['client_id'])
        order = Order.objects.create(
            manager=request.user, client=client, 
            order_number=new_order_number,
            invoice_number=new_order_number,
            destination=data.get('destination', ''),
            transition_point=data.get('transition_point', ''),
            shipper=data.get('shipper', ''),
            shipper_address=data.get('shipper_address', ''),
            consignee=data.get('consignee', ''),
            consignee_address=data.get('consignee_address', ''),
            transport=transport,
            loaded_weight=data.get('loaded_weight') or None,
            currency=data.get('currency', 'USD')
        )

        # --- ИСПРАВЛЕНИЕ ЗДЕСЬ: Расчет total_quantity ---
        total_quantity = 0
        for prod_data in products_input_data:
            product = get_object_or_404(Product, id=prod_data['product_id'])
            order_product = OrderProduct.objects.create(
                order=order, product=product, 
                quantity=int(prod_data['quantity']),
                gross_weight=Decimal(prod_data['gross_weight'])
            )
            total_quantity += order_product.quantity
        
        order.calculate_totals()

        car_info_str = "/".join(filter(None, [order.transport.car_number, order.transport.trailer_number])) if order.transport else 'N/A'

        product_name_str = "Сборный груз"
        if order.order_products.count() == 1:
            product_name_str = order.order_products.first().product.product_name

        data_for_gsheet = {
            'date': order.created_at.strftime('%Y-%m-%d'),
            'order_number': order.order_number,
            'car_info': car_info_str,
            'product_name': product_name_str,
            'gross_weight': float(order.gross_weight) / 1000,
            'net_weight': float(order.netto_weight) / 1000,
            'total_amount': f"{order.total_amount}"
        }

        # 2. Вызываем функцию отправки
        append_to_google_sheet(data_for_gsheet)
        
        # --- Генерация Excel-файла ---
        # Теперь инвойс будет создаваться с уже перерасчитанными весами
        # ... (весь ваш код генерации Excel, который использует total_quantity)

        

        # --- НАЧАЛО ЛОГИКИ СОЗДАНИЯ EXCEL ИНВОЙСА ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Стили
        bold_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Ширина колонок
        column_widths = {'B': 5, 'C': 30, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 18}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Заполнение шапки документа
        ws['C1'] = 'INVOICE / PACKING LIST'; ws['C1'].font = bold_font; ws['C1'].alignment = center_align
        ws['I2'] = 'INVOICE №'; ws['J2'] = order.invoice_number
        ws['I3'] = 'DATE'; ws['J3'] = datetime.now().strftime('%d.%m.%Y')
        
        ws['B4'] = 'ПОЛУЧАТЕЛЬ'; ws['C4'] = order.consignee
        ws['C5'] = order.consignee_address
        ws['B7'] = 'ОТПРАВИТЕЛЬ'; ws['C7'] = order.shipper
        ws['C8'] = order.shipper_address

        car_display_number = 'N/A'
        if order.transport:
            parts = [order.transport.car_number, order.transport.trailer_number]
            car_display_number = "/".join(filter(None, parts)) or 'N/A'
        
        ws['B10'] = 'КОНТРАКТ'; ws['C10'] = order.order_products.first().product.contract if order.order_products.exists() else 'N/A'
        ws['B11'] = 'ПУНКТ НАЗНАЧЕНИЯ'; ws['C11'] = order.destination
        ws['B12'] = 'АВТОТРАНСПОРТ'; ws['C12'] = car_display_number

        # Заголовки таблицы
        headers = ['№', 'Наименование товара', 'Код ТН ВЭД', 'Марка', 'Товарный знак', 'Страна происхождения', 'Производитель', 'Кол-во грузовых мест', 'Вес нетто(кг)', 'Вес брутто (кг)', 'Стоимость ({order.currency})']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=15, column=col); cell.value = header; cell.font = bold_font; cell.alignment = center_align; cell.border = border

        # Заполнение строк с товарами
        row_num = 16
        for i, op in enumerate(order.order_products.all(), start=1):
            product_data = [i, op.product.product_name, op.product.code, op.product.product_mark, op.product.product_mark, op.product.product_country, op.product.product_mark, op.quantity, op.netto_weight, op.gross_weight, op.amount]
            for col, value in enumerate(product_data, start=2):
                cell = ws.cell(row=row_num, column=col); cell.value = value; cell.alignment = center_align; cell.border = border
            row_num += 1
            
        # Строка с итогами
        ws[f'B{row_num}'] = 'Total:'; ws[f'B{row_num}'].font = bold_font; ws[f'B{row_num}'].alignment = right_align; ws.merge_cells(f'B{row_num}:H{row_num}')
        totals_data = [total_quantity, order.netto_weight, order.gross_weight, order.total_amount]
        for i, val in enumerate(totals_data, start=9):
            cell = ws.cell(row=row_num, column=i); cell.value = val; cell.font = bold_font; cell.alignment = center_align; cell.border = border

        # Сохранение файла в память
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Создание объекта OrderFile и сохранение файла
        order_file = OrderFile(order=order)
        order_file.file.save(f"SLK_invoice_{order.invoice_number}.xlsx", output, save=True)

        # --- КОНЕЦ ЛОГИКИ СОЗДАНИЯ EXCEL ИНВОЙСА ---

        return JsonResponse({'success': True, 'order_id': order.id, 'redirect_url': f'/orders/{order.id}/'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@login_required
def profile_page(request):
    if request.method == 'POST':
        # Определяем, какая из двух форм была отправлена
        if 'update_profile' in request.POST:
            profile_form = UserProfileUpdateForm(request.POST, instance=request.user)
            password_form = PasswordChangeForm(request.user) # Создаем пустую форму пароля
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, 'Ваш профиль успешно обновлен.')
                return redirect('profile_page')

        elif 'change_password' in request.POST:
            profile_form = UserProfileUpdateForm(instance=request.user) # Создаем пустую форму профиля
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                # Важный шаг, чтобы пользователь не вышел из системы после смены пароля
                update_session_auth_hash(request, user)
                messages.success(request, 'Ваш пароль был успешно изменен.')
                return redirect('profile_page')
    else:
        # При GET-запросе просто создаем пустые формы, заполненные данными пользователя
        profile_form = UserProfileUpdateForm(instance=request.user)
        password_form = PasswordChangeForm(request.user)

    return render(request, 'profile.html', {
        'profile_form': profile_form,
        'password_form': password_form
    })

def root_redirect_view(request):
    if request.user.is_authenticated:
        return redirect('order_list_page') 
    else:
        return redirect('login_page')